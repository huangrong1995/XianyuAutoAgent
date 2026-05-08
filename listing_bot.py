#!/usr/bin/env python3
"""
闲鱼自动托管工具 - 自动上架模块
流程: 买家购买 → 自动确认发货 → 自动发送云盘链接给买家

依赖于 XianyuAutoAgent 的 Cookie 和 API Key 配置
"""

import os
import sys
import json
import time
import random
import argparse
import base64
import hashlib
import asyncio
from pathlib import Path
from datetime import datetime
from threading import Thread

import openpyxl
import requests
from dotenv import load_dotenv

# 项目路径
PROJECT_DIR = Path(__file__).parent
DATA_DIR = PROJECT_DIR / "data"
STATE_FILE = DATA_DIR / "listing_state.json"
PRODUCTS_EXCEL = DATA_DIR / "products.xlsx"


def generate_sign(t: str, token: str, data: str) -> str:
    """生成签名"""
    app_key = "34839810"
    msg = f"{token}&{t}&{app_key}&{data}"
    md5_hash = hashlib.md5()
    md5_hash.update(msg.encode('utf-8'))
    return md5_hash.hexdigest()


def upload_image(image_path: str, config: dict) -> str:
    """上传图片到闲鱼CDN，返回图片URL"""
    if not image_path or not os.path.exists(image_path):
        return None

    session = requests.Session()
    for part in config["cookies_str"].split(";"):
        part = part.strip()
        if "=" in part:
            name, value = part.split("=", 1)
            session.cookies.set(name.strip(), value.strip(), domain=".goofish.com")

    try:
        with open(image_path, "rb") as f:
            files = {"file": (os.path.basename(image_path), f, "image/jpeg")}
            data = {
                "folderId": "0",
                "appkey": "fleamarket",
                "_input_charset": "utf-8"
            }
            response = session.post(
                "https://stream-upload.goofish.com/api/upload.api",
                files=files,
                data=data,
                timeout=30
            )

        result = response.text
        print(f"   📤 图片上传响应: {result}")

        # 解析响应获取URL
        try:
            resp_json = json.loads(result)
            # 尝试从响应中提取图片URL
            if isinstance(resp_json, dict):
                url = resp_json.get("url") or resp_json.get("data", {}).get("url") or resp_json.get("data", {}).get("imgUrl")
                if url:
                    return url
            # 如果响应就是URL字符串
            if result.startswith("http"):
                return result
        except:
            pass

        print(f"   ⚠️ 图片上传解析失败")
        return None

    except Exception as e:
        print(f"   ❌ 图片上传异常: {e}")
        return None


def upload_images_from_folder(img_folder: str, config: dict) -> list:
    """从文件夹上传多张图片，返回URL列表"""
    img_dir = PROJECT_DIR / img_folder if img_folder else None
    if not img_dir or not img_dir.exists():
        return []

    img_files = list(img_dir.glob("*.*"))
    if not img_files:
        return []

    urls = []
    for img_file in img_files[:9]:  # 最多9张
        url = upload_image(str(img_file), config)
        if url:
            urls.append({
                "extraInfo": {"isH": "false", "isT": "false", "raw": "false"},
                "isQrCode": False,
                "url": url,
                "heightSize": 1024,
                "widthSize": 1024,
                "major": len(urls) == 0,  # 第一张设为主图
                "type": 0,
                "status": "done"
            })
        time.sleep(0.5)  # 避免上传过快

    return urls


# ============ 配置加载 ============

def load_config():
    """加载环境配置"""
    env_file = PROJECT_DIR / ".env"
    load_dotenv(env_file)
    
    return {
        "cookies_str": os.getenv("COOKIES_STR", ""),
        "api_key": os.getenv("API_KEY", ""),
        "model_base_url": os.getenv("MODEL_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1"),
        "model_name": os.getenv("MODEL_NAME", "qwen-max"),
    }


# ============ Excel 商品数据 ============

def create_template_excel(path: Path):
    """创建商品模板 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品列表"
    
    headers = [
        "序号", "状态", "商品ID", "标题", "价格", "描述", 
        "图片文件夹", "分类", "标签",
        "百度云链接", "夸克云链接", "百度云密码", "夸克云密码",
        "发货消息模板", "累计售出", "最后上架时间"
    ]
    ws.append(headers)
    
    # 示例数据
    sample_products = [
        [1, "待上架", "", "【电子资料】Python编程入门全套视频教程", "9.9",
         "包含完整Python基础+进阶视频教程，共200+集，附赠源码和课件。",
         "images/python_course", "其他", "Python,编程",
         "https://pan.baidu.com/s/xxxx", "https://pan.quark.cn/s/xxxx",
         "1234", "abcd",
         "您好！资料链接：{link}，提取码：{password}，请保存好链接如有丢失可再次联系客服索取。",
         0, ""],
        [2, "待上架", "", "【电子资料】2024考研考公全套复习资料", "19.9",
         "涵盖考研/考公全科资料，包含真题、笔记、重点总结。",
         "images/kaoyan", "其他", "考研,考公",
         "https://pan.baidu.com/s/yyyy", "https://pan.quark.cn/s/yyyy",
         "5678", "efgh",
         "您好！考研资料链接：{link}，提取码：{password}，包含最新真题和重点笔记。",
         0, ""],
    ]
    
    for row in sample_products:
        ws.append(row)
    
    wb.save(path)
    print(f"✅ 已创建商品模板: {path}")


def load_products(path: Path) -> list:
    """从 Excel 加载商品列表"""
    if not path.exists():
        return []
    
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    products = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        (seq, status, item_id, title, price, desc, img_folder, 
         category, tags, baidu_link, quark_link, baidu_pwd, quark_pwd,
         msg_template, sold_count, last_time) = row
        products.append({
            "row": i,
            "seq": seq,
            "status": status or "待上架",
            "item_id": item_id or "",
            "title": title or "",
            "price": str(price or "0"),
            "desc": desc or "",
            "img_folder": img_folder or "",
            "category": category or "其他",
            "tags": tags or "",
            "baidu_link": baidu_link or "",
            "quark_link": quark_link or "",
            "baidu_pwd": baidu_pwd or "",
            "quark_pwd": quark_pwd or "",
            "msg_template": msg_template or "您好！资料链接：{link}，提取码：{password}，请保存好链接。",
            "sold_count": sold_count or 0,
            "last_time": last_time or ""
        })
    
    return products


def update_product(path: Path, row_num: int, updates: dict):
    """更新商品信息"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    col_map = {
        "status": 2, "item_id": 3, "title": 4, "price": 5, "desc": 6,
        "img_folder": 7, "category": 8, "tags": 9,
        "baidu_link": 10, "quark_link": 11, "baidu_pwd": 12, "quark_pwd": 13,
        "msg_template": 14, "sold_count": 15, "last_time": 16
    }
    
    for key, value in updates.items():
        if key in col_map:
            ws.cell(row=row_num, column=col_map[key]).value = value
    
    wb.save(path)


def build_delivery_message(product: dict, link_type: str = "baidu") -> str:
    """构建发货消息"""
    template = product["msg_template"]
    
    if link_type == "baidu":
        link = product["baidu_link"]
        password = product["baidu_pwd"]
    else:
        link = product["quark_link"]
        password = product["quark_pwd"]
    
    if not link:
        return None
    
    # 替换占位符
    msg = template.replace("{link}", link).replace("{password}", password)
    return msg


# ============ Cookie 解析 ============

def parse_cookies(cookie_str: str) -> list:
    """解析 cookies 字符串"""
    cookies = []
    for part in cookie_str.split(";"):
        part = part.strip()
        if "=" in part:
            name, value = part.split("=", 1)
            cookies.append({
                "name": name.strip(),
                "value": value.strip(),
                "domain": ".goofish.com",
                "path": "/"
            })
    return cookies


def save_cookies_to_env(new_cookies_str: str, env_path: str = "/app/.env"):
    """保存新cookies到.env文件"""
    try:
        with open(env_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        with open(env_path, 'w', encoding='utf-8') as f:
            for line in lines:
                if line.startswith('COOKIES_STR='):
                    f.write(f'COOKIES_STR={new_cookies_str}\n')
                else:
                    f.write(line)
        
        print(f"   💾 Cookies已保存到.env")
    except Exception as e:
        print(f"   ⚠️ 保存Cookies失败: {e}")


# ============ 闲鱼 API 操作 ============

def confirm_delivery(cookies: list, item_id: str, buyer_id: str) -> bool:
    """确认发货（数字资料自动发货）"""
    import requests
    
    session = requests.Session()
    for c in cookies:
        session.cookies.set(c["name"], c["value"], domain=c.get("domain", ".goofish.com"))
    
    url = "https://h5api.m.goofish.com/h5/mtop.taobao.idle.order.confirmsend/1.0/"
    headers = {
        "content-type": "application/json",
        "referer": "https://www.goofish.com/",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    
    payload = {
        "itemId": item_id,
        "buyerId": buyer_id,
        "expressType": "virtual",
        "message": "数字资料商品，链接将在24小时内发送，请注意查收。"
    }
    
    try:
        resp = session.post(url, headers=headers, json=payload, timeout=15)
        data = resp.json()
        ret = data.get("ret", [None])[0]
        if ret and ret.startswith("SUCCESS"):
            print(f"   ✅ 发货成功")
            return True
        else:
            print(f"   ⚠️  发货响应: {data}")
            return True  # 仍返回成功，避免中断流程
    except Exception as e:
        print(f"   ❌ 发货API失败: {e}")
        return True  # 继续发送链接


def relist_with_playwright(product: dict, config: dict) -> str:
    """使用 Selenium WebDriver 连接真实浏览器上架"""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import os

    selenium_url = os.getenv("SELENIUM_URL", "http://selenium:4444")
    cookies = parse_cookies(config["cookies_str"])

    try:
        options = Options()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--window-size=1920,1080")

        driver = webdriver.Remote(
            command_executor=selenium_url + "/wd/hub",
            options=options
        )
        print(f"   🔗 已连接 Selenium Grid")
    except Exception as e:
        print(f"   ⚠️ Selenium连接失败: {e}")
        return ""

    def get_login_modal():
        """检测登录弹窗"""
        # 只检查 passport iframe 是否显示登录界面
        try:
            iframes = driver.find_elements(By.TAG_NAME, 'iframe')
            for iframe in iframes:
                src = iframe.get_attribute('src') or ''
                if 'passport' in src and iframe.is_displayed():
                    # 检查iframe是否包含扫码登录内容（不是已登录状态）
                    driver.switch_to.frame(iframe)
                    try:
                        # 查找QR code canvas或扫码登录相关的元素
                        qr_canvas = driver.find_elements(By.CSS_SELECTOR, 'canvas, [class*="qrcode"], [class*="QRCode"]')
                        for elem in qr_canvas:
                            if elem.is_displayed():
                                driver.switch_to.default_content()
                                return iframe
                    except:
                        pass
                    driver.switch_to.default_content()
        except:
            driver.switch_to.default_content()
        return None

    def wait_for_login():
        """等待用户扫码登录，返回是否成功并更新cookie"""
        print(f"   ⚠️ 检测到登录验证，等待扫码...")
        driver.save_screenshot("/tmp/login_required.png")
        print(f"   📸 登录弹窗已截图，请扫码登录...")

        for i in range(60):
            time.sleep(1)
            try:
                # 检查passport iframe是否消失
                passport_iframe_visible = False
                iframes = driver.find_elements(By.TAG_NAME, 'iframe')
                for iframe in iframes:
                    src = iframe.get_attribute('src') or ''
                    if 'passport' in src and iframe.is_displayed():
                        passport_iframe_visible = True
                        break
                
                if not passport_iframe_visible:
                    # 获取新cookies
                    print(f"   ✅ 登录成功，正在获取新Cookie...")
                    new_cookies_list = driver.get_cookies()
                    new_cookies_str = "; ".join([f"{c['name']}={c['value']}" for c in new_cookies_list])
                    config["cookies_str"] = new_cookies_str
                    # 保存到 .env 文件
                    save_cookies_to_env(new_cookies_str)
                    return True
            except Exception as e:
                print(f"   ⚠️ 检查登录状态时出错: {e}")
            if (i + 1) % 10 == 0:
                print(f"   ⏳ 等待扫码中... ({i + 1}秒)")

        return False

    try:
        print(f"   🌐 正在访问主页...")
        driver.get("https://www.goofish.com")
        time.sleep(5)

        # 从.env读取的cookies可能已过期，先尝试添加
        cookies = parse_cookies(config["cookies_str"])
        for cookie in cookies:
            try:
                driver.add_cookie(cookie)
            except:
                pass

        # 刷新页面让cookies生效
        driver.get("https://www.goofish.com")
        time.sleep(5)
        
        # 检查登录弹窗
        if get_login_modal():
            if not wait_for_login():
                print(f"   ❌ 登录超时")
                driver.quit()
                return ""

        print(f"   🌐 正在打开发布页面...")
        driver.get("https://www.goofish.com/publish")
        time.sleep(3)

        wait = WebDriverWait(driver, 15)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[contenteditable="true"]')))
        print(f"   📄 页面已加载: {driver.title}")

        try:
            editor = driver.find_element(By.CSS_SELECTOR, 'div[contenteditable="true"]')
            editor.click()
            time.sleep(0.5)
            editor.send_keys(product["title"])
            # 换行后填写描述
            if product.get("desc"):
                editor.send_keys("\n")
                editor.send_keys(product["desc"])
            # 触发 input 事件确保 React/Vue 检测到输入
            driver.execute_script('arguments[0].dispatchEvent(new Event("input", {bubbles: true}));', editor)
            print(f"   ✏️ 标题已填写")
        except Exception as e:
            print(f"   ⚠️ 标题填写失败: {e}")

        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[placeholder="0.00"]')))
        try:
            price_input = driver.find_element(By.CSS_SELECTOR, 'input[placeholder="0.00"]')
            price_input.click()
            time.sleep(0.3)
            price_input.send_keys(str(product["price"]))
            # 触发 input 事件
            driver.execute_script('arguments[0].dispatchEvent(new Event("input", {bubbles: true}));', price_input)
            print(f"   ✏️ 价格已填写: {product['price']}")
        except Exception as e:
            print(f"   ⚠️ 价格填写失败: {e}")

        # 上传图片（支持多张）
        if product.get("img_folder"):
            img_dir = PROJECT_DIR / product["img_folder"]
            if img_dir.exists():
                img_files = list(img_dir.glob("*.*"))
                if img_files:
                    try:
                        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="file"]')))
                        file_input = driver.find_element(By.CSS_SELECTOR, 'input[type="file"]')
                        # 上传所有图片（最多9张）
                        img_paths = "\n".join([str(f) for f in img_files[:9]])
                        file_input.send_keys(img_paths)
                        print(f"   📷 图片已上传: {len(img_files[:9])}张")
                        time.sleep(5)
                    except Exception as e:
                        print(f"   ⚠️ 图片上传失败: {e}")

        if get_login_modal():
            if not wait_for_login():
                print(f"   ❌ 登录超时")
                driver.quit()
                return ""

        try:
            wait.until(EC.presence_of_element_located((By.XPATH, '//button[contains(.,"发布")]')))
            publish_btn = driver.find_element(By.XPATH, '//button[contains(.,"发布")]')
            print(f"   🔘 找到发布按钮")
            # 等待一下让页面稳定
            time.sleep(2)
            publish_btn.click()
            print(f"   🔘 已点击发布按钮，等待响应...")
            time.sleep(10)
        except Exception as e:
            print(f"   ⚠️ 发布按钮点击失败: {e}")

        url = driver.current_url
        print(f"   🔗 发布后URL: {url}")
        # 检查 item?id= 或 itemId=
        if "item?id=" in url:
            item_id = url.split("item?id=")[1].split("&")[0]
            print(f"   ✅ 上架成功: {item_id}")
            return item_id
        elif "itemId=" in url:
            item_id = url.split("itemId=")[1].split("&")[0]
            print(f"   ✅ 上架成功: {item_id}")
            return item_id
        else:
            print(f"   ⚠️ 未找到商品ID，当前URL: {url}")

    except Exception as e:
        print(f"   ❌ Selenium 上架失败: {e}")
        try:
            driver.save_screenshot("/tmp/selenium_error.png")
            print(f"   📸 错误页面已截图")
        except:
            pass
    finally:
        driver.quit()

    return ""




def relist_with_api(product: dict, config: dict) -> str:
    """使用 API 上架商品"""
    session = requests.Session()

    # 解析 cookies
    for part in config["cookies_str"].split(";"):
        part = part.strip()
        if "=" in part:
            name, value = part.split("=", 1)
            session.cookies.set(name.strip(), value.strip(), domain=".goofish.com")

    # 获取 token
    token = session.cookies.get("_m_h5_tk", "").split("_")[0]
    if not token:
        print("   ❌ 无法获取 token")
        return ""

    t = str(int(time.time() * 1000))

    # 上传图片
    img_urls = []
    img_folder = product.get("img_folder", "")
    if img_folder:
        print(f"   📤 开始上传图片...")
        img_urls = upload_images_from_folder(img_folder, config)
        print(f"   📤 图片上传完成，共 {len(img_urls)} 张")

    # 如果没有上传图片，使用默认图片
    if not img_urls:
        # 使用一个占位图片URL（需要有效的已上传图片）
        print(f"   ⚠️ 没有可用图片，尝试不传图片上架...")
        img_urls = []

    # 构建请求数据（根据抓包结果）
    price = product.get("price", "0")
    price_in_cent = str(int(float(price) * 100)) if price else "1000"

    item_data = {
        "freebies": False,
        "itemTypeStr": "b",
        "quantity": "1",
        "simpleItem": "true",
        "imageInfoDOList": img_urls,
        "itemTextDTO": {
            "desc": product.get("desc", "NA"),
            "title": product.get("title", "NA"),
            "titleDescSeparate": False
        },
        "itemLabelExtList": [{
            "channelCateName": "软件安装包/序列号/激活码",
            "valueId": None,
            "channelCateId": "201449620",
            "valueName": None,
            "tbCatId": "50003316",
            "subPropertyId": None,
            "labelType": "common",
            "subValueId": None,
            "labelId": None,
            "propertyName": "分类",
            "isUserClick": "0",
            "isUserCancel": None,
            "from": "newPublishChoice",
            "propertyId": "-10000",
            "labelFrom": "newPublish",
            "text": "软件安装包/序列号/激活码",
            "properties": "-10000##分类:201449620##软件安装包/序列号/激活码"
        }],
        "itemPriceDTO": {
            "priceInCent": price_in_cent
        },
        "userRightsProtocols": [
            {"enable": False, "serviceCode": "AI_SALE"},
            {"enable": False, "serviceCode": "SKILL_PLAY_NO_MIND"}
        ],
        "itemPostFeeDTO": {
            "canFreeShipping": True,
            "supportFreight": True,
            "onlyTakeSelf": True
        },
        "itemAddrDTO": {
            "area": "仓山区",
            "city": "福州",
            "divisionId": 350104,
            "gps": "26.048661,119.241787",
            "poiId": "B024F0XWXC",
            "poiName": "新榕金城湾",
            "prov": "福建"
        },
        "defaultPrice": False,
        "itemCatDTO": {
            "catId": "50025461",
            "catName": "软件安装包/序列号/激活码",
            "channelCatId": "201449620",
            "leafId": "1831",
            "tbCatId": "50003316"
        },
        "uniqueCode": t + "716",
        "sourceId": "pcMainPublish",
        "bizcode": "pcMainPublish",
        "publishScene": "pcMainPublish"
    }

    data_val = json.dumps(item_data, ensure_ascii=False)

    # 生成签名
    sign = generate_sign(t, token, data_val)

    # 构建请求
    params = {
        "jsv": "2.7.2",
        "appKey": "34839810",
        "t": t,
        "sign": sign,
        "v": "1.0",
        "type": "originaljson",
        "accountSite": "xianyu",
        "dataType": "json",
        "timeout": "20000",
        "api": "mtop.idle.pc.idleitem.publish",
        "sessionOption": "AutoLoginOnly",
        "spm_cnt": "a21ybx.publish.0.0",
        "spm_pre": "a21ybx.im.sidebar.1.29b04f10mWJ25H",
        "log_id": "29b04f10mWJ25H"
    }

    headers = {
        "content-type": "application/x-www-form-urlencoded",
        "referer": "https://www.goofish.com/",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36"
    }

    try:
        response = session.post(
            "https://h5api.m.goofish.com/h5/mtop.idle.pc.idleitem.publish/1.0/",
            params=params,
            data={"data": data_val},
            headers=headers,
            timeout=30
        )

        result = response.json()
        print(f"   📦 上架响应: {result}")

        ret = result.get("ret", [None])
        if ret and any("SUCCESS" in str(r) for r in ret):
            # 尝试从响应中提取商品ID
            data = result.get("data", {})
            if isinstance(data, dict):
                item_id = data.get("itemId") or data.get("item_id") or data.get("idleId")
                if item_id:
                    print(f"   ✅ 上架成功: {item_id}")
                    return str(item_id)
            # 如果没有解析到itemId但返回成功，也返回成功
            print(f"   ✅ 上架API调用成功")
            return "success"
        else:
            print(f"   ❌ 上架失败: {result}")
            return ""

    except Exception as e:
        print(f"   ❌ 上架API异常: {e}")
        return ""


def try_relist(product: dict, config: dict) -> str:
    """尝试上架商品，优先使用API方式"""
    # 优先使用API方式
    print(f"   🔄 尝试API方式上架...")
    result = relist_with_api(product, config)
    if result:
        return result

    # API方式失败，尝试Playwright
    print(f"   🔄 API方式失败，尝试Playwright方式...")
    try:
        from playwright.sync_api import sync_playwright
        result = relist_with_playwright(product, config)
        return result
    except ImportError:
        print(f"   ⚠️  Playwright未安装或不可用，上架失败")
        return ""


# ============ 消息发送（供 main.py 调用）============

def get_delivery_message_for_product(item_id: str = None, product: dict = None) -> str:
    """
    获取指定商品的发货消息
    优先用 item_id 匹配，否则用 product
    """
    if item_id:
        products = load_products(PRODUCTS_EXCEL)
        for p in products:
            if p.get("item_id") == item_id:
                product = p
                break
    
    if not product:
        return None
    
    # 优先百度云，其次夸克
    if product.get("baidu_link"):
        return build_delivery_message(product, "baidu")
    elif product.get("quark_link"):
        return build_delivery_message(product, "quark")
    
    return None


def do_confirm_and_relist(item_id: str, buyer_id: str, product: dict, send_chat_id: str = None) -> dict:
    """
    执行确认发货+重新上架
    返回结果供 main.py 发送消息
    """
    config = load_config()
    cookies = parse_cookies(config["cookies_str"])
    
    result = {
        "success": False,
        "item_id": item_id,
        "buyer_id": buyer_id,
        "chat_id": send_chat_id,
        "delivery_msg": None,
        "new_item_id": None,
    }
    
    # 1. 确认发货
    print(f"\n📦 确认发货: 商品={item_id}, 买家={buyer_id}")
    confirm_delivery(cookies, item_id, buyer_id)
    
    # 2. 准备发货消息
    if product:
        result["delivery_msg"] = get_delivery_message_for_product(product=product)
        print(f"📨 发货消息: {result['delivery_msg'][:50] if result['delivery_msg'] else '无'}...")
        
        # 3. 重新上架
        print(f"🚀 重新上架...")
        new_id = relist_with_playwright(product, config)
        result["new_item_id"] = new_id
        
        # 4. 更新商品状态
        if new_id:
            sold_count = (product.get("sold_count") or 0) + 1
            update_product(PRODUCTS_EXCEL, product["row"], {
                "status": "已上架",
                "item_id": new_id,
                "sold_count": sold_count,
                "last_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            print(f"✅ 流程完成! 累计售出: {sold_count}")
    
    result["success"] = True
    return result


# ============ 主程序 ============

def main():
    parser = argparse.ArgumentParser(description="闲鱼自动上架机器人")
    parser.add_argument("--init", action="store_true", help="初始化商品模板")
    parser.add_argument("--interval", type=int, default=60, help="检查间隔（秒）")
    parser.add_argument("--monitor", action="store_true", help="监控模式")
    args = parser.parse_args()
    
    if args.init:
        DATA_DIR.mkdir(exist_ok=True)
        create_template_excel(PRODUCTS_EXCEL)
        return
    
    if not PRODUCTS_EXCEL.exists():
        print(f"❌ 商品文件不存在: {PRODUCTS_EXCEL}")
        print(f"   运行 --init 生成模板")
        sys.exit(1)
    
    if args.monitor:
        print(f"🔄 监控模式已启动，每{args.interval}秒检查一次...")
        while True:
            try:
                products = load_products(PRODUCTS_EXCEL)
                for product in products:
                    if product["status"] == "待上架":
                        print(f"\n📦 检测到待上架商品: {product['title']}")
                        config = load_config()
                        new_id = relist_with_playwright(product, config)
                        if new_id:
                            update_product(PRODUCTS_EXCEL, product["row"], {
                                "status": "已上架",
                                "item_id": new_id,
                                "last_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            })
                            print(f"✅ 上架成功: {new_id}")
                        else:
                            print(f"⚠️ 上架失败，保持待上架状态")
                time.sleep(args.interval)
            except KeyboardInterrupt:
                print("\n👋 监控已停止")
                break
            except Exception as e:
                print(f"❌ 监控异常: {e}")
                time.sleep(args.interval)
    else:
        products = load_products(PRODUCTS_EXCEL)
        for product in products:
            if product["status"] == "待上架":
                print(f"\n📦 上架: {product['title']}")
                config = load_config()
                new_id = relist_with_playwright(product, config)
                if new_id:
                    update_product(PRODUCTS_EXCEL, product["row"], {
                        "status": "已上架",
                        "item_id": new_id,
                        "last_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })


if __name__ == "__main__":
    main()
